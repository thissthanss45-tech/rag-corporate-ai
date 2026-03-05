from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any

import docx2txt
import fitz
from celery import Task
from fastembed import SparseTextEmbedding, TextEmbedding
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from striprtf.striprtf import rtf_to_text

from app.celery_app import celery_app
from app.config import settings
from app.services.qdrant_service import QdrantService

logger = logging.getLogger(__name__)

_embedder: TextEmbedding | None = None
_sparse_embedder: SparseTextEmbedding | None = None

FATAL_PARSE_ERRORS = (ValueError, fitz.FileDataError, fitz.EmptyFileError, UnicodeDecodeError, InvalidFileException)


class TextSplitter:
    def __init__(self, chunk_size: int, chunk_overlap: int) -> None:
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def split_text(self, text: str) -> list[str]:
        chunks: list[str] = []
        start = 0
        text_len = len(text)

        while start < text_len:
            end = start + self.chunk_size
            if end < text_len:
                last_newline = text.rfind("\n", start, end)
                if last_newline != -1 and last_newline > start + self.chunk_size // 2:
                    end = last_newline + 1
                else:
                    last_space = text.rfind(" ", start, end)
                    if last_space != -1 and last_space > start + self.chunk_size // 2:
                        end = last_space + 1

            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)

            next_start = end - self.chunk_overlap
            start = next_start if next_start > start else end

        return chunks


def get_embedder() -> TextEmbedding:
    global _embedder
    if _embedder is None:
        logger.info("🧠 loading embedding model", extra={"model": settings.EMBEDDING_MODEL_NAME})
        _embedder = TextEmbedding(model_name=settings.EMBEDDING_MODEL_NAME)
    return _embedder


def get_sparse_embedder() -> SparseTextEmbedding:
    global _sparse_embedder
    if _sparse_embedder is None:
        logger.info("🔍 loading sparse model", extra={"model": settings.SPARSE_EMBEDDING_MODEL_NAME})
        _sparse_embedder = SparseTextEmbedding(model_name=settings.SPARSE_EMBEDDING_MODEL_NAME)
    return _sparse_embedder


def to_sparse_parts(vector: Any) -> tuple[list[int], list[float]]:
    indices = getattr(vector, "indices", None)
    values = getattr(vector, "values", None)

    if indices is None and isinstance(vector, dict):
        indices = vector.get("indices")
        values = vector.get("values")

    if indices is None or values is None:
        raise ValueError("Unsupported sparse vector format")

    return [int(item) for item in indices], [float(item) for item in values]


def extract_text_from_pdf(file_path: str) -> str:
    text_parts: list[str] = []
    with fitz.open(file_path) as doc:
        for page in doc:
            text_parts.append(page.get_text("text"))
    return "\n".join(text_parts)


def extract_text_from_docx(file_path: str) -> str:
    return docx2txt.process(file_path) or ""


def extract_text_from_txt(file_path: str) -> str:
    payload = Path(file_path).read_bytes()
    for encoding in ("utf-8", "cp1251", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("txt", payload, 0, min(1, len(payload)), "Unsupported TXT encoding")


def extract_text_from_xlsx(file_path: str) -> str:
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    sheet_texts: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(item).strip() for item in row if item is not None and str(item).strip()]
            if values:
                rows.append("\t".join(values))
        if rows:
            sheet_texts.append(f"[SHEET: {sheet.title}]\n" + "\n".join(rows))
    workbook.close()
    return "\n\n".join(sheet_texts)


def extract_text_from_rtf(file_path: str) -> str:
    payload = Path(file_path).read_bytes()
    for encoding in ("utf-8", "cp1251", "latin-1"):
        try:
            raw = payload.decode(encoding)
            return rtf_to_text(raw)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("rtf", payload, 0, min(1, len(payload)), "Unsupported RTF encoding")


def extract_text(file_path: str) -> str:
    suffix = Path(file_path).suffix.lower()
    if suffix == ".pdf":
        return extract_text_from_pdf(file_path)
    if suffix == ".docx":
        return extract_text_from_docx(file_path)
    if suffix == ".txt":
        return extract_text_from_txt(file_path)
    if suffix == ".xlsx":
        return extract_text_from_xlsx(file_path)
    if suffix == ".rtf":
        return extract_text_from_rtf(file_path)
    raise ValueError(f"Unsupported file type: {suffix}")


@celery_app.task(bind=True, max_retries=3, name="app.tasks.process_document_task")
def process_document_task(self: Task, file_path: str, file_name: str) -> dict[str, int | str]:
    logger.info("⚙️ start document processing", extra={"task_id": self.request.id, "file_name": file_name})

    try:
        raw_text = extract_text(file_path)
        if not raw_text.strip():
            logger.warning("🗑 empty text extracted", extra={"task_id": self.request.id, "file_name": file_name})
            return {"status": "skipped", "chunks": 0, "points": 0}

        splitter = TextSplitter(chunk_size=settings.CHUNK_SIZE, chunk_overlap=settings.CHUNK_OVERLAP)
        chunks = splitter.split_text(raw_text)
        if not chunks:
            logger.warning("🗑 no chunks created", extra={"task_id": self.request.id, "file_name": file_name})
            return {"status": "skipped", "chunks": 0, "points": 0}

        logger.info("🧠 vectorizing chunks", extra={"task_id": self.request.id, "chunks": len(chunks)})
        embedder = get_embedder()
        vectors = [list(map(float, vector)) for vector in embedder.embed(chunks)]

        sparse_embedder = get_sparse_embedder()
        sparse_raw = list(sparse_embedder.embed(chunks))
        sparse_vectors = [to_sparse_parts(item) for item in sparse_raw]
        vector_size = len(vectors[0]) if vectors else settings.VECTOR_SIZE

        qdrant_service = QdrantService()
        qdrant_service.ensure_collection(vector_size=vector_size)
        saved_points = qdrant_service.upsert_chunks(
            dense_vectors=vectors,
            sparse_vectors=sparse_vectors,
            file_name=file_name,
            chunks=chunks,
        )

        logger.info(
            "✅ document processed",
            extra={
                "task_id": self.request.id,
                "file_name": file_name,
                "chunks": len(chunks),
                "points": saved_points,
            },
        )
        return {"status": "completed", "chunks": len(chunks), "points": saved_points}

    except FATAL_PARSE_ERRORS as exc:
        logger.error(
            "🗑 document rejected as invalid",
            extra={"task_id": self.request.id, "file_name": file_name, "error": str(exc)},
        )
        return {"status": "failed", "chunks": 0, "points": 0}
    except Exception as exc:
        logger.warning(
            "↩️ task retry scheduled",
            extra={"task_id": self.request.id, "file_name": file_name, "error": str(exc)},
        )
        raise self.retry(exc=exc, countdown=30)
    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(
                    "🗑 file physically deleted from disk",
                    extra={"task_id": self.request.id, "file_name": file_name, "file_path": file_path},
                )
        except OSError as exc:
            logger.warning(
                "⚠️ failed to delete file from disk",
                extra={"task_id": self.request.id, "file_name": file_name, "file_path": file_path, "error": str(exc)},
            )
